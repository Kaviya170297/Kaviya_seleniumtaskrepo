from openpyxl import load_workbook

def get_row_count(file, sheet_name):
    workbook = load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row

def get_column_count(file, sheet_name):
    workbook = load_workbook(file)      #Open workbook
    sheet = workbook[sheet_name]                   #choose sheet
    return sheet.max_column

def read_data(file, sheet_name, row_num, column_num):
    workbook = load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=column_num).value


def write_data(file, sheet_name, row_num, column_num, data):
    workbook = load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=column_num).value = data
    workbook.save(file)
    workbook.close()